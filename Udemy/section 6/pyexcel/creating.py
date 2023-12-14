# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

# Facilitar ao avisar que uma variavel é de determinado tipo (nesse caso worksheet)
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Local de arquivo do arquivo main.py
ROOT_FOLDER = Path(__file__).parent
# arquivo excel
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Instancia um Workbook
workbook = Workbook()

# Criar a planilha
sheet_name = 'Planilha'
workbook.create_sheet(sheet_name)

# Selecionar planilha
# Avisar ao vscode o tipo do worksheet
worksheet: Worksheet = workbook[sheet_name]

# Deletar uma planilha
workbook.remove(workbook['Sheet'])


# Criando cabeçalhos (linha, coluna, valor)
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome      idade  nota
    ['Noki',    20,    10],
    ['Jhonson', 22,    8],
    ['Mike',    24,    9],
    ['Mari',    19,    9.5],
]

# Maneira mais complicada, porem IMPORTANTE
# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

# Maneira mais simples e usada
for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)