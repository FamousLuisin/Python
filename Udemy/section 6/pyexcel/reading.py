from pathlib import Path

# Facilitar ao avisar que uma variavel é de determinado tipo (nesse caso worksheet)
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

# Local de arquivo do arquivo main.py
ROOT_FOLDER = Path(__file__).parent

# arquivo excel
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Instancia um Workbook
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Criar a planilha
sheet_name = 'Planilha'

# Selecionar planilha
# Avisar ao vscode o tipo do worksheet
worksheet: Worksheet = workbook[sheet_name]

# min_row > começar por tal linha
# Ta criando um interavel com cada linha
for row in worksheet.iter_rows(min_row=2):
    # Recebe uma linha
    # Pega a coluna dessa linha
    for col in row:
        # Pega o valor a partir da coluna.value
        print(col.value, end='\t')
        
        # Caso o col.value seja == Jhonson, ele saberá q esta na linha certa
        if col.value == 'Jhonson':
            # Assim, usando o col.row tem como descobri a linha daquele valor
            # E sabendo qual coluna colocar é so inserir no segundo parametro
            # No terceiro o valor a ser trocado
            worksheet.cell(col.row, 2, 20)
    print()

# Alterar um valor no worbook['Local_na_planilha']
# worksheet['B3'].value = 32
print(worksheet['B3'].value)


workbook.save(WORKBOOK_PATH)